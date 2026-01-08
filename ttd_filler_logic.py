import pandas as pd
from openpyxl import load_workbook
import re

# ---------------- HELPERS ----------------

def clean_mobile(mobile):
    if pd.isna(mobile):
        return ""
    digits = re.sub(r"\D", "", str(mobile))
    if len(digits) > 10 and digits.startswith("91"):
        digits = digits[2:]
    return digits if len(digits) == 10 else ""

def split_address(addr):
    if pd.isna(addr) or not addr:
        return "", "", ""

    parts = [p.strip() for p in str(addr).split(",") if p.strip()]

    # remove last 3 → city, state, pincode
    if len(parts) > 3:
        parts = parts[:-3]

    line1 = parts[0] if len(parts) > 0 else ""
    line2 = parts[1] if len(parts) > 1 else ""
    line3 = ", ".join(parts[2:]) if len(parts) > 2 else line2

    return line1, line2, line3


# ---------------- VOLUMETRIC ----------------

def load_volumetric_tables(path):
    df = pd.read_excel(path, header=None)

    def block(r1, r2, c1, c2):
        b = df.iloc[r1:r2, c1:c2].copy()
        b.columns = ["Quantity", "L", "B", "H", "Weight"]
        b["Quantity"] = pd.to_numeric(b["Quantity"], errors="coerce")
        b = b.dropna(subset=["Quantity"])
        b["Quantity"] = b["Quantity"].astype(int)
        return b.set_index("Quantity")

    return {
        "calendar": block(2, 22, 0, 5),
        "ttc": block(2, 27, 6, 11),
        "big_diary": block(28, 48, 0, 5),
        "small_diary": block(28, 53, 6, 11),
    }

def get_dimensions(vol, category, qty):
    qty = int(qty)
    c = str(category).lower()

    if "calendar" in c and "table" not in c:
        table = vol["calendar"]
    elif "table" in c:
        table = vol["ttc"]
    elif "big" in c:
        table = vol["big_diary"]
    elif "small" in c:
        table = vol["small_diary"]
    else:
        return None, None, None

    if qty in table.index:
        r = table.loc[qty]
    else:
        r = table[table.index <= qty].iloc[-1]

    return int(r["L"]), int(r["B"]), int(r["H"])


# ---------------- MAIN LOGIC ----------------

def generate_output(orders_path, postal_path, template_path, volumetric_path, output_path):

    # ---- POSTAL FILE ----
    postal = pd.read_excel(postal_path, header=3)

    tr_col = postal.columns[1]
    name_col = postal.columns[2]
    addr_col = postal.columns[3]
    city_col = postal.columns[4]
    pin_col = postal.columns[5]
    mobile_col = postal.columns[6]
    qty_col = postal.columns[7]
    weight_col = postal.columns[8]
    barcode_col = postal.columns[9]

    postal = postal[[tr_col, name_col, addr_col, city_col, pin_col,
                      mobile_col, qty_col, weight_col, barcode_col]].copy()

    postal["__TR"] = postal[tr_col].astype(str).str.strip()
    postal["Receiver name"] = postal[name_col]
    postal["Full Address"] = postal[addr_col]
    postal["Receiver city"] = postal[city_col]
    postal["Receiver pincode"] = pd.to_numeric(postal[pin_col], errors="coerce").fillna(0).astype(int)
    postal["Receiver mobile"] = postal[mobile_col].apply(clean_mobile)
    postal["Quantity"] = pd.to_numeric(postal[qty_col], errors="coerce").fillna(1).astype(int)
    postal["Physical weight in grams"] = pd.to_numeric(postal[weight_col], errors="coerce").fillna(0).astype(int)
    postal["Barcode"] = postal[barcode_col].astype(str).str.strip()

    postal = postal[postal["Receiver pincode"].between(100000, 999999)].copy()

    # ---- ORDERS FILE ----
    orders = pd.read_excel(orders_path, sheet_name="Publications_Report")
    orders["__TR"] = orders["Booking No"].astype(str).str.strip()

    orders = orders[[
        "__TR",
        "State",
        "Address"
    ]].drop_duplicates("__TR")

    # ---- MERGE ----
    merged = postal.merge(orders, on="__TR", how="left")
    merged["State"] = merged["State"].fillna("Tamil Nadu")

    # 🔴 ONLY NEW FIX (AS REQUESTED)
    merged["Full Address"] = merged["Full Address"].fillna(merged["Address"])

    # ---- LOAD TEMPLATE ----
    wb = load_workbook(template_path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    defaults = [c.value for c in ws[2]]

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    vol = load_volumetric_tables(volumetric_path)

    serial = 1

    for idx, row in merged.iterrows():
        r = 2 + idx
        L, B, H = get_dimensions(vol, row["Category"], row["Quantity"])

        for i, header in enumerate(headers):
            cell = ws.cell(row=r, column=i + 1)
            h = "" if header is None else header.lower()

            if "serial" in h:
                cell.value = serial
            elif "barcode" in h:
                cell.value = row["Barcode"]
            elif "physical weight" in h:
                cell.value = row["Physical weight in grams"]
            elif "receiver city" in h:
                cell.value = row["Receiver city"]
            elif "receiver pincode" in h:
                cell.value = row["Receiver pincode"]
            elif "receiver name" in h:
                cell.value = row["Receiver name"]
            elif "receiver mobile" in h:
                cell.value = row["Receiver mobile"]
            elif "add line 1" in h:
                cell.value = split_address(row["Full Address"])[0]
            elif "add line 2" in h:
                cell.value = split_address(row["Full Address"])[1]
            elif "add line 3" in h:
                cell.value = split_address(row["Full Address"])[2]
            elif "length" in h:
                cell.value = L
            elif "breadth" in h or "diameter" in h:
                cell.value = B
            elif "height" in h:
                cell.value = H
            elif "receiver state" in h:
                cell.value = row["State"]
            else:
                cell.value = defaults[i]

            if "sender mobile" in h:
                cell.value = 1234567890
            if "sender state" in h:
                cell.value = "Andhra Pradesh"
            if "sender add line 1" in h:
                cell.value = "SALES WING OF PUBLICATIONS"
            if "sender add line 2" in h:
                cell.value = "TTD PRESS COMPOUND"
            if "sender add line 3" in h:
                cell.value = "Tirupati-517507"

        serial += 1

    wb.save(output_path)
    return serial - 1
